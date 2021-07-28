# TODO: extract features of all videos
import copy
from operator import mod
import os
import time
from collections import Counter, deque

import cv2
import face_alignment
import numpy as np
import torch
import torch.nn.functional as F
import torchvision
from imutils import face_utils
from PIL import Image
from scipy.spatial import distance as dist
from scipy.stats import entropy
from torchvision import transforms, utils

import util
from configs.image_cfg import _C as cfg
from dataset import make_data_loader
from dataset.transforms import build_transforms
from model.graph_net import Graph_Net
from model.overall_net import Net

# ----- 20210728 ----- #
from pathlib import Path
import openpyxl as xl

def main():
    schizo_videos_dir = "/myHDD/chihyuan_data/schizo_data"
    extensions = ("**/*.mp4", "**/*.MP4")

    annotated_videos = get_annotated_videos()

    count_annotated = 0
    for extension in extensions:
        for video in Path(schizo_videos_dir).glob(extension):
            video_name = video.name.split('.')[0]
            if video_name in annotated_videos:
                video_path = video.absolute()
                count_annotated += 1
                print(count_annotated, video_path)
                infer(video_path=video_path, video_name=video_name)
            else:
                print(f"Not Annotated: {video}")
    print(f"Total number of annotated videos: {count_annotated}")


def get_annotated_videos():
    workbook = xl.load_workbook("/home/chihyuan/ntu_schizo/data/Rating_20210420.xlsx")
    sheet = workbook["工作表1"]
    video_names = []

    for col in range(2, sheet.max_column + 1):
        cell_video_name = sheet.cell(1, col)
        cell_ng = sheet.cell(2, col)
        if "(NG影片)" not in str(cell_ng.value):
            # print(col, cell_video_name.value)
            video_names.append(str(cell_video_name.value))
    
    return video_names
# -------------------- #

# ----- 20210728 ----- #
def infer(video_path="./test.mp4", video_name="test"):
# -------------------- #
    os.environ['CUDA_VISIBLE_DEVICES'] = cfg.MODEL.DEVICE_ID
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    torch.backends.cudnn.benchmark = True

    # train_loader, test_loader, num_classes, label_template = make_data_loader(cfg)
    label_template = ['Angry', 'Disgust', 'Fear', 'Happy', 'Neutral', 'Sad', 'Surprise']
    pos_template   = ['Happy', 'Surprise']
    neg_template   = ['Angry', 'Disgust', 'Fear', 'Sad']
    neutral_class  = 'Neutral'

    num_classes    = len(label_template)
    # define and load model
    model_path = cfg.MODEL.SAVE_WEIGHT_PATH+'.pth'
    model = Net(cfg, num_classes).to(device)
    model.load_state_dict(torch.load(model_path))
    model.eval()

    # face detector
    fa = face_alignment.FaceAlignment(face_alignment.LandmarksType._2D, flip_input=False, device='cuda')

    (lBegin, lEnd) = face_utils.FACIAL_LANDMARKS_IDXS["right_eyebrow"]
    (rBegin, rEnd) = face_utils.FACIAL_LANDMARKS_IDXS["left_eyebrow"]
    
    f_transform = build_transforms(cfg, is_train=False, is_face=True)
    c_transform = build_transforms(cfg, is_train=False, is_face=False)

    # video path
    # ----- 20210728 ----- #
    # video_path = './test.mp4'
    # -------------------- #
    cap = cv2.VideoCapture(video_path)
    f_stack = deque([])
    c_stack = deque([])
    seq_pred = deque([])
    points   = deque([])
    seq_arousal_val = deque([])
    long_seq_pred = deque([])
    long_seq_prob = deque([])
    long_seq_size = 64
    window_size = 16
    # ----- 20210722 ----- #
    video_features = []  # list of all the features along time of a whole video
    # -------------------- #
    if (cap.isOpened()== False): 
        print("Error opening video stream or file")
    while(cap.isOpened()):
        ret, frame = cap.read()
        if ret == True:
            if cv2.waitKey(25) & 0xFF == ord('q'):
                break
            frame = cv2.resize(frame,None,fx=0.5,fy=0.5,interpolation=cv2.INTER_CUBIC)  
            # cv2.imwrite('./test.png', frame)
            # print(Image.open('./test.png').size, torchvision.io.read_image('./test.png').shape, frame.shape)
            # assert i == -1
            # cv2.imshow('Frame', frame) 
            # cv2.putText(frame, label_table[np.argmax(emotion_count)], (max_face_info[0], int(max_face_info[2] - 5)), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 4)
            
            # ----- 20170721 ----- #
            try:
                face, context, landmark = detFace(fa, frame)
            except:
                print(f"No faces were detected (Error at line 81 from inference.py)")
                continue
            # -------------------- #
            

            collectBatch(face, f_transform, f_stack, window_size)
            collectBatch(context, c_transform, c_stack, window_size)

            leyebrow = landmark[lBegin:lEnd]
            reyebrow = landmark[rBegin:rEnd]
            distq = eye_brow_distance(points,window_size,leyebrow[-1],reyebrow[0])
            seq_arousal_val.append(normalize_values(points,distq))
            
            if len(seq_arousal_val) > window_size:
                seq_arousal_val.popleft()

            if len(f_stack) == window_size:

                with torch.no_grad():
                    f_tensor = torch.stack(list(copy.deepcopy(f_stack))).to(device)
                    c_tensor = torch.stack(list(copy.deepcopy(c_stack))).to(device)

                    # ----- 20210722 ----- #
                    # output = model(f_tensor, c_tensor).mean(0)
                    output, features = model(f_tensor, c_tensor)
                    features = features.mean(0)
                    output = output.mean(0)
                    # ----- 20210722 ----- #
                    # model_output.append(output.cpu().tolist())
                    # -------------------- #
                    video_features.append(features.cpu().tolist())
                    # -------------------- #
                    pred_class = output.argmax()
                    pred_label = label_template[pred_class]
                    arousal_label = ''
                    arousal_value = np.mean(seq_arousal_val)
                    # if arousal_value > 0.75:
                    #     arousal_label = 'high arousal'
                    # else:
                    #     arousal_label = 'low  arousal'
                    
                    # print(label_template[pred_class], np.mean(seq_arousal_val))
                    det_disorder = "normal"
                    long_seq_pred.append(pred_label)
                    long_seq_prob.append(output.cpu().numpy())
                    text_color = (0, 255, 0)
                    if len(long_seq_pred) > long_seq_size:
                        long_seq_pred.popleft()
                        long_seq_prob.popleft()

                        mean_pred = np.mean(long_seq_prob, 0)
                        mean_prob = prob(mean_pred)

                        # count prediction
                        count_window_pred = Counter(long_seq_pred)
                        count_window_pred_list = []
                        for val in count_window_pred:
                            count_window_pred_list.append(count_window_pred[val])
                        count_window_pred_list = np.array(count_window_pred_list, dtype=np.float) / long_seq_size
                        max_count_pred    = count_window_pred.most_common(1)[0]
                        
                        disorder = False
                        if entropy(count_window_pred_list) > 0.9 or arousal_value > 0.75:
                            disorder = True
                        
                        if disorder:
                            det_disorder = "abnormal"
                            text_color = (0, 0, 255)
                        # assert i == -1
                        
                        # if max_count_pred[1] > (long_seq_size // 2) and max_count_pred[0] != pred_label:
                        #     det_disorder = "abnormal"
                        #     text_color = (0, 0, 255)
                        # print(count_window_pred, max_count_pred)
                        # assert i == -1
                    
                    cv2.putText(frame, pred_label, 
                                (20, 30), 
                                cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 4)
                    
                    cv2.putText(frame, 'arousal => {:.4f}'.format(arousal_value), 
                                (20, 70), 
                                cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 4)
                    
                    cv2.putText(frame, det_disorder, 
                                (20, 110), 
                                cv2.FONT_HERSHEY_SIMPLEX, 1, text_color, 4)

                    # show_face = cv2.cvtColor(face, cv2.COLOR_RGB2BGR)
                    cv2.imshow('frame', frame)
        else: 
            break

    # ----- 20210728 ----- #
    torch.save(video_features, "output/video_features/" + video_name + ".pth")
    # -------------------- #

    cap.release()
    cv2.destroyAllWindows()

def prob(log_output):
    return np.exp(log_output)/sum(np.exp(log_output))

def eye_brow_distance(points, size, leye, reye):
    distq = dist.euclidean(leye,reye)
    points.append(int(distq))
    if len(points) > size:
        points.popleft()
    return distq

def normalize_values(points, disp):
    # print(points)
    normalized_value = abs(disp - np.min(points))/abs(np.max(points) - np.min(points))
    arousal_value = np.exp(-(normalized_value))
    return arousal_value

def collectBatch(image, transform, stack, size):
    tensor = transform(Image.fromarray(image))
    stack.append(tensor)
    if len(stack) > size:
        stack.popleft()

def detFace(model, image):
    process_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    w, h, c = process_image.shape
    preds = model.get_landmarks(process_image)
    if preds != None and len(preds) > 0:
        max_face_area = 0
        max_face_info = [0, 0, 0, 0]
        max_landmark = None
        try:
            for i, landmark in enumerate(preds):
                        
                x1 = np.min(landmark[:, 0])
                y1 = np.min(landmark[:, 1])
                x2 = np.max(landmark[:, 0])
                y2 = np.max(landmark[:, 1])

                box_len = int((float(max(x2 - x1, y2 - y1)) / 2.) * 1.3)
                center_x = int(float(x1 + x2) / 2.)
                center_y = int(float(y1 + y2) / 2.)
                
                if box_len > max_face_area:
                    max_face_area = box_len
                    max_face_info = [center_x, center_y]
                    max_landmark = landmark

            sx = max_face_info[0] - max_face_area
            ex = max_face_info[0] + max_face_area
            sy = max_face_info[1] - max_face_area
            ey = max_face_info[1] + max_face_area
            face = process_image[sx:ex, sy:ey, :]
            wf, hf, cf = face.shape
            result = np.full((max_face_area * 2,max_face_area * 2, c), (0,0,0), dtype=np.uint8)
            xx = (max_face_area * 2 - wf) // 2
            yy = (max_face_area * 2 - hf) // 2
            result[xx:xx+wf, yy:yy+hf] = process_image[sx:ex, sy:ey, :]
            
            return result, process_image, landmark
        except Exception as e: 
            print(e)
            mid_w = w // 2
            mid_h = h // 2
            box_len = min(mid_h, mid_w) // 2
            return process_image[(mid_w - box_len):(mid_w + box_len), (mid_h - box_len):(mid_h + box_len), :], process_image, None

if __name__ == "__main__":
    main()
